from flask import Flask, render_template, request, redirect, url_for, flash
import sqlite3
import pandas as pd
from openpyxl import load_workbook
import os

app = Flask(__name__)
app.secret_key = "secreto"

def obter_dados():
    conexao = sqlite3.connect('dados.db')
    consulta = "SELECT * FROM tabela_dados"
    df_consulta = pd.read_sql_query(consulta, conexao)
    conexao.close()
    return df_consulta

@app.route('/')
def index():
    return render_template('index.html', df_consulta=obter_dados().to_html(classes='table table-striped table-bordered table-sm', index=False))

@app.route('/consulta', methods=['POST'])
def consultar_dados():
    dado_procurado = request.form['entrada_consulta']
    consulta = f"""
                 SELECT *
                 FROM tabela_dados
                 WHERE Sistema LIKE '%{dado_procurado}%' OR
                     EQUIPAMENTO LIKE '%{dado_procurado}%'  OR 
                     PROTOCOLO LIKE '%{dado_procurado}%'    OR 
                     MODELO LIKE '%{dado_procurado}%'       OR
                     FABRICANTE LIKE '%{dado_procurado}%'   OR
                     MEIO_FISICO LIKE  '%{dado_procurado}%' OR
                     DESCRIÇÃO  LIKE   '%{dado_procurado}%' OR
                     SUBSISTEMA LIKE '%{dado_procurado}%'
                 """
    conexao = sqlite3.connect('dados.db')
    df_consulta = pd.read_sql_query(consulta, conexao)
    conexao.close()
    return render_template('index.html', df_consulta=df_consulta.to_html(classes='table table-striped table-bordered table-sm', index=False))

@app.route('/adicionar_equipamento', methods=['POST'])
def adicionar_equipamento():
    # Adapte a lógica para adicionar equipamento
    flash('Equipamento adicionado com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/remover_equipamento/<int:item_id>')
def remover_equipamento(item_id):
    # Adapte a lógica para remover equipamento
    flash('Equipamento removido com sucesso!', 'success')
    return redirect(url_for('index'))

@app.route('/gerar_excel', methods=['POST'])
def gerar_excel_selecionados():
    lista = obter_dados().values.tolist()
    # Adapte a lógica para gerar o Excel
    workbook = load_workbook('LM-T-08-07-01-1299_6-R11-001_Rev_A.xlsx')
    sheet = workbook['ListaEquipamentos']
    linha_de_insercao = 11

    for i, sublist in enumerate(lista, start=1):
        sublist.insert(0, i)
        for col, valor in zip(range(1, len(sublist) + 1), sublist):
            sheet.cell(row=linha_de_insercao+i, column=col, value=valor)

    nome_arquivo = 'Nova_lista_pequipamentos.xlsx'
    if os.path.exists(nome_arquivo):
        nome_base, extensao = os.path.splitext(nome_arquivo)
        contador = 1
        while True:
            novo_nome_arquivo = f"{nome_base}_{contador}{extensao}"
            if not os.path.exists(novo_nome_arquivo):
                break
            contador += 1
        nome_arquivo = novo_nome_arquivo

    workbook.save(nome_arquivo)
    flash('Documento Gerado!', 'success')
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
