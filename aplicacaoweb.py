from flask import Flask, render_template, request, send_file
import sqlite3
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
lista = []


app = Flask(__name__)


# Variável global para armazenar os resultados da consulta
df_consulta = pd.DataFrame()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/consultar', methods=['POST'])
def consultar_dados():
    global df_consulta
    dado_procurado = request.form.get('consulta')
    consulta = f"""
               SELECT *
               FROM tabela_dados
               WHERE Sistema LIKE '%{dado_procurado}%' OR
                    EQUIPAMENTO LIKE '%{dado_procurado}%'  OR
                    PROTOCOLO LIKE '%{dado_procurado}%'    OR
                    MODELO LIKE '%{dado_procurado}%'       OR
                    FABRICANTE LIKE '%{dado_procurado}%'   OR
                    MEIO_FISICO LIKE  '%{dado_procurado}%' OR
                    DESCRIÇÃO  LIKE   '%{dado_procurado}%' OR
                    SUBSISTEMA LIKE '%{dado_procurado}%'
               """
    conexao = sqlite3.connect('dados.db')
    df_consulta = pd.read_sql_query(consulta, conexao)
    preencher_tabela(df_consulta)
    conexao.close()
    return render_template('index.html', tabela=df_consulta.to_html(escape=False), status="")


@app.route('/gerar_excel', methods=['POST'])
def gerar_excel_selecionados():

    global df_consulta
    linhas_selecionadas = request.form.getlist('linhas_selecionadas[]')
    if linhas_selecionadas:
        linhas_selecionadas = [int(idx) for idx in linhas_selecionadas if int(idx) < len(df_consulta)]  # Garantindo que os índices não ultrapassem o tamanho do DataFrame
        if linhas_selecionadas:
            df_selecionados = df_consulta.iloc[linhas_selecionadas]
            print(df_selecionados)
            df_selecionados.drop(columns=['Selecionar'], inplace=True)  # Excluir a coluna "Selecionar"
            print(df_selecionados)
            nome_arquivo = 'meu_dataframe.xlsx'
            lista = df_selecionados.values.tolist()
            for indice, sublist in enumerate(lista, start=1):
                sublist.insert(0, indice)
            # Carregando o arquivo Excel existente
            workbook = load_workbook('C:/Project_Python/InterfaceWeb/lista_equipamentos_tt36.xlsx')

            # Selecionando a planilha na qual você deseja adicionar a linha
            sheet = workbook['ListaEquipamentos']
            linha_de_insercao = 11

            nova_linha = lista  # Substitua pelos valores desejados
            for i in range(len(nova_linha)):
                for col, valor in zip(range(1, len(nova_linha[i]) + 1), nova_linha[i]):
                    sheet.cell(row=linha_de_insercao + i, column=col, value=valor)

            nome_arquivo = 'Nova_lista_equipamentos.xlsx'


            workbook.save(nome_arquivo)
            return send_file(nome_arquivo, as_attachment=True)
    return render_template('index.html', tabela=df_consulta.to_html(escape=False), status="Nenhuma linha selecionada.")




def preencher_tabela(df):
    df['Selecionar'] = '<input type="checkbox" name="linhas_selecionadas[]" value="' + df.index.astype(str) + '">'
    tabela = df.to_html(escape=False)
    return tabela


def criar_excel(df, nome_arquivo):
    wb = Workbook()
    ws = wb.active
    print(ws)
    for r_idx, row in enumerate(df.values, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)
            print(r_idx,c_idx,value)
    # Inserir o cabeçalho
    for col, column_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col, value=column_name)

    # Estilizar o cabeçalho
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    # Ajustar a largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    wb.save(nome_arquivo)

if __name__ == '__main__':
    app.run(debug=True)